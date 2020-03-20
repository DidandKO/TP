from openpyxl import load_workbook
import matplotlib as mpl
import matplotlib.pyplot as plt
import math
import csv


# метод для получения значений тока на осях робота при нулевой нагрузке
# возвращает набор точек (название операции/среднее значение тока за 3 итерации)
def cargo_axis_pointer(_cargo, _axis):
    try:
        # cargo_axis = []
        # среднее значение тока за 3 итеррации по осям и грузам
        vb_cargo = {0: 3, 1: 15, 2: 27, 3: 39, 4: 51}
        vb_axis = {1: 'S', 2: 'T', 3: 'U', 4: 'V', 5: 'W', 6: 'X'}
        wb = load_workbook("D:\Python\PyProjs\TP\media\DATATP.xlsx", data_only=True)
        print(wb.sheetnames)
        table = wb['Таблица']
        operation = []
        electricity = []
        for i in range(12):
            _adressX = 'R' + str(vb_cargo.get(_cargo)+i)
            _adressY = str(vb_axis.get(_axis)) + str(vb_cargo.get(_cargo)+i)
            operation.append(table[_adressX].value)
            electricity.append(table[_adressY].value)
            print(str(table[_adressX].value) + '-' + str(table[_adressY].value))
        #     cargo_axis.append([table[_adressX].value, table[_adressY].value])
        # print(cargo_axis)

        # FILENAME = "D:\Python\PyProjs\TP\media\my.csv"
        # with open(FILENAME, "w", newline="") as file:
        #     writer = csv.writer(file)
        #     writer.writerows(cargo_axis)

        dpi = 80
        fig = plt.figure(dpi=dpi, figsize=(1024 / dpi, 768 / dpi))
        mpl.rcParams.update({'font.size': 10})

        plt.axis([20, 60, -12, 12])

        plt.title('Axis Electricity')
        plt.xlabel('Operation')
        plt.ylabel('Electricity')

        plt.plot(operation, electricity, color='blue', linestyle='solid',
                 label=_cargo)

        plt.legend(loc='upper right')
        fig.savefig('D:\Python\PyProjs\TP\schedule\static\schedule\images\scheda.png')

    except FileNotFoundError:
        print("no")



